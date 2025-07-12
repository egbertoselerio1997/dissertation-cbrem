import os
import qsdsan as qs
import pandas as pd
import numpy as np
from joblib import Parallel, delayed
import warnings
from tqdm import tqdm
import openpyxl
import gc

# Suppress the specific pkg_resources UserWarning
warnings.filterwarnings("ignore", category=UserWarning, message=".*pkg_resources is deprecated.*")

def run_single_simulation(i, input_df, vary_inputs):
    """
    Runs a single simulation instance. This function is designed to be parallelized.
    """
    try:
        # Each parallel worker needs its own QSDsan environment setup
        cmps = qs.processes.create_asm2d_cmps()
        qs.set_thermo(cmps)
        
        current_inputs = {}
        for var in input_df.index:
            is_randomizable = input_df.loc[var, 'randomizable'] == 1
            if vary_inputs and is_randomizable:
                low = input_df.loc[var, 'min']
                high = input_df.loc[var, 'max']
                current_inputs[var] = np.random.uniform(low, high)
            else:
                current_inputs[var] = input_df.loc[var, 'min']

        influent_concentrations = {
            key.replace('inf_', ''): value
            for key, value in current_inputs.items() if key.startswith('inf_')
        }
        ws = qs.WasteStream(f'influent_ws_{i}')
        ws.set_flow_by_concentration(flow_tot=current_inputs['flow_rate'], concentrations=influent_concentrations, units=('m3/d', 'mg/L'))

        asm2d_model = qs.processes.ASM2d()

        V_max = ws.F_vol * (current_inputs['HRT'] / 24)

        reactor = qs.sanunits.CSTR(
            ID=f'R_{i}',
            ins=ws.copy(),
            V_max=V_max,
            aeration=current_inputs['DO_setpoint'],
            suspended_growth_model=asm2d_model,
            DO_ID='S_O2'
        )
        
        sys = qs.System(f'sys_{i}', path=(reactor,))
        sys.simulate(t_span=(0, 50), method='BDF', state_reset_hook='reset_cache')
        
        outputs = reactor.results(with_units=False)

        effluent_stream = reactor.outs[0]
        component_ids = reactor.components.IDs
        
        effluent_concs = effluent_stream.conc
        
        effluent_data = {}
        
        for idx, comp_id in enumerate(component_ids):
            effluent_data[f'Target_Effluent_{comp_id} (mg/L)'] = effluent_concs[idx]

        run_data = {
            'simulation_number': i + 1,
            **current_inputs,
            **outputs,
            **effluent_data
        }
        
        return run_data
    except Exception:
        return None
    finally:
        # Clean up QSDsan objects created in this worker process to prevent memory leaks
        # The library holds objects in registries, which must be cleared manually.
        qs.main_flowsheet.clear()

# Main simulation function
def run_simulation():
    """
    Main function to drive the simulation based on user inputs and Excel data.
    """
    data_filepath = os.path.join('data', 'data.xlsx')

    print(f"Loading input variables from {data_filepath}...")
    try:
        input_df = pd.read_excel(data_filepath, sheet_name='input_config', index_col=0)
    except FileNotFoundError:
        print(f"Error: '{data_filepath}' not found. Please create it or run the script again to generate a template.")
        return
    except ValueError:
        print(f"Error: 'input_config' sheet not found in '{data_filepath}'.")
        return

    while True:
        try:
            num_simulations = int(input("Enter the number of simulations to run: "))
            if num_simulations > 0:
                break
            print("Please enter a positive number.")
        except ValueError:
            print("Invalid input. Please enter an integer.")

    vary_inputs = False
    if num_simulations > 1:
        while True:
            choice = input("Vary inputs randomly? (y/n): ").lower()
            if choice in ('y', 'n'):
                vary_inputs = (choice == 'y')
                break
            print("Invalid choice. Please enter 'y' or 'n'.")

    while True:
        try:
            batch_size = int(input("Enter the batch size for processing (e.g., 100): "))
            if batch_size > 0:
                break
            print("Please enter a positive number.")
        except ValueError:
            print("Invalid input. Please enter an integer.")

    total_successful = 0
    
    try:
        book = openpyxl.load_workbook(data_filepath)
        sheets_to_delete = ['results_statistics']
        for sheet_name in sheets_to_delete:
            if sheet_name in book.sheetnames:
                del book[sheet_name]
        book.save(data_filepath)
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"Warning: Could not clear previous result sheets in '{data_filepath}'. Error: {e}")

    print(f"\nStarting {num_simulations} simulations in batches of {batch_size}...")
    for i in tqdm(range(0, num_simulations, batch_size), desc="Processing Batches"):
        batch_end = min(i + batch_size, num_simulations)
        
        with Parallel(n_jobs=-1) as parallel:
            batch_results = parallel(
                delayed(run_single_simulation)(sim_num, input_df, vary_inputs)
                for sim_num in range(i, batch_end)
            )

        successful_batch_results = [r for r in batch_results if r is not None]
        total_successful += len(successful_batch_results)

        if not successful_batch_results:
            continue

        results_df = pd.DataFrame(successful_batch_results)
        
        input_cols = ['simulation_number'] + input_df.index.to_list()
        input_results_df = results_df.reindex(columns=input_cols)
        
        output_cols = ['simulation_number'] + [col for col in results_df.columns if col not in input_df.index.to_list() and col != 'simulation_number']
        output_results_df = results_df.reindex(columns=output_cols)

        try:
            with pd.ExcelWriter(data_filepath, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                is_new_input_sheet = 'results_input' not in writer.book.sheetnames
                input_start_row = 0 if is_new_input_sheet else writer.book['results_input'].max_row
                input_results_df.to_excel(writer, sheet_name='results_input', index=False, header=is_new_input_sheet, startrow=input_start_row)

                is_new_output_sheet = 'results_output' not in writer.book.sheetnames
                output_start_row = 0 if is_new_output_sheet else writer.book['results_output'].max_row
                output_results_df.to_excel(writer, sheet_name='results_output', index=False, header=is_new_output_sheet, startrow=output_start_row)
        except Exception as e:
            print(f"\nAn error occurred while saving batch results: {e}")
            break
        
        # Explicitly clear memory
        del batch_results
        del successful_batch_results
        del results_df
        del input_results_df
        del output_results_df
        gc.collect()

    print(f"\n{total_successful}/{num_simulations} simulations completed successfully.")

    if total_successful == 0:
        print("No results to save or analyze.")
        return

    print(f"All results have been incrementally saved to '{data_filepath}'.")
    
    if total_successful > 1:
        print("Calculating final statistics from the saved results...")
        try:
            final_output_df = pd.read_excel(data_filepath, sheet_name='results_output')
            stats_df = final_output_df.drop(columns=['simulation_number']).describe().T
            
            with pd.ExcelWriter(data_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                stats_df.to_excel(writer, sheet_name='results_statistics')
            
            print(f"Final statistics saved to sheet 'results_statistics' in '{data_filepath}'.")
        except Exception as e:
            print(f"An error occurred during final statistics calculation: {e}")

if __name__ == '__main__':
    run_simulation()