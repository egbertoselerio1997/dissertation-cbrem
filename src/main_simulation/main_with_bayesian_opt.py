import os
import qsdsan as qs
import pandas as pd
import numpy as np
from joblib import Parallel, delayed
import warnings
from functools import partial
from tqdm import tqdm
import openpyxl
import gc

# Suppress the specific pkg_resources UserWarning
warnings.filterwarnings("ignore", category=UserWarning, message=".*pkg_resources is deprecated.*")

try:
    from scipy.stats import norm
except ImportError:
    print("SciPy not found. Please install it using: pip install scipy")
    exit()

try:
    import optuna
except ImportError:
    print("Optuna not found. Please install it using: pip install optuna")
    exit()

try:
    from rich.live import Live
    from rich.panel import Panel
except ImportError:
    print("Rich not found. Please install it using: pip install rich")
    exit()


class OptunaProgressCallback:
    """
    Optuna callback to provide live visual feedback and handle early stopping based on patience.
    """
    def __init__(self, patience, max_calls):
        self.patience = patience
        self.max_calls = max_calls
        self.best_value = -np.inf
        self.n_no_improvement = 0
        self.live = None

    def __call__(self, study: optuna.study.Study, trial: optuna.trial.FrozenTrial):
        current_value = study.best_value if study.best_value is not None else -np.inf
        
        if current_value > self.best_value:
            self.best_value = current_value
            self.n_no_improvement = 0
        else:
            self.n_no_improvement += 1

        header_text = f"Trial: {trial.number + 1}/{self.max_calls} | "
        header_text += f"Best Avg Std: {self.best_value:.4f} | "
        header_text += f"Patience: {self.n_no_improvement}/{self.patience}"
        
        status_panel = Panel(header_text, title="[bold green]Bayesian Optimization Status[/bold green]", border_style="green")

        if self.live:
            self.live.update(status_panel)

        if self.n_no_improvement >= self.patience:
            study.stop()

def run_single_simulation(i, input_df, vary_inputs, probability_val=None):
    """
    Runs a single simulation instance. This function is designed to be parallelized.
    """
    try:
        # Each parallel worker needs its own QSDsan environment setup
        cmps = qs.processes.create_asm2d_cmps()
        qs.set_thermo(cmps)
        
        current_inputs = {}
        for var in input_df.index:
            is_randomizable = input_df.loc[var, 'randomizable'] == 1
            if vary_inputs and is_randomizable:
                if probability_val is None:
                    raise ValueError("Probability value must be provided for random variation.")
                
                mean_val = input_df.loc[var, 'default']
                std_dev = input_df.loc[var, 'std. dev.']
                
                # Calculate z-score for the given probability
                z = norm.ppf(1 - (1 - probability_val) / 2)
                
                # Estimate min and max for uniform distribution
                low = mean_val - z * std_dev
                high = mean_val + z * std_dev
                
                # Apply non-negativity constraint to the lower bound
                low = max(0, low)
                
                # Generate a random number from a uniform distribution
                current_inputs[var] = np.random.uniform(low, high)
            else:
                current_inputs[var] = input_df.loc[var, 'default']

        influent_concentrations = {
            key.replace('inf_', ''): value
            for key, value in current_inputs.items() if key.startswith('inf_')
        }
        ws = qs.WasteStream(f'influent_ws_{i}')
        ws.set_flow_by_concentration(flow_tot=current_inputs['flow_rate'], concentrations=influent_concentrations, units=('m3/d', 'mg/L'))

        asm2d_model = qs.processes.ASM2d()

        V_max = ws.F_vol * (current_inputs['HRT'] / 24)

        reactor = qs.sanunits.CSTR(
            ID=f'R_{i}',
            ins=ws.copy(),
            V_max=V_max,
            aeration=current_inputs['DO_setpoint'],
            suspended_growth_model=asm2d_model,
            DO_ID='S_O2'
        )
        
        sys = qs.System(f'sys_{i}', path=(reactor,))
        sys.simulate(t_span=(0, 50), method='BDF', state_reset_hook='reset_cache')
        
        outputs = reactor.results(with_units=False)

        effluent_stream = reactor.outs[0]
        component_ids = reactor.components.IDs
        
        effluent_concs = effluent_stream.conc
        
        effluent_data = {}
        
        for idx, comp_id in enumerate(component_ids):
            effluent_data[f'Effluent_{comp_id} (mg/L)'] = effluent_concs[idx]

        run_data = {
            'simulation_number': i + 1,
            **current_inputs,
            **outputs,
            **effluent_data
        }
        
        return run_data
    except Exception:
        return None

def objective_function(trial, base_input_df, optim_vars, num_simulations, probability_val):
    """
    Objective function for Optuna. It runs a set of stochastic
    simulations and returns the average standard deviation of the outputs.
    """
    input_df = base_input_df.copy()
    input_df['default'] = input_df['baseline']
    
    # Suggest new parameters for this trial
    for var in optim_vars:
        baseline_val = base_input_df.loc[var, 'baseline']
        std_dev = base_input_df.loc[var, 'std. dev.']
        low = max(0, baseline_val - 3 * std_dev)
        high = baseline_val + 3 * std_dev
        input_df.loc[var, 'default'] = trial.suggest_float(var, low, high)

    all_results = Parallel(n_jobs=-1)(
        delayed(run_single_simulation)(i, input_df, vary_inputs=True, probability_val=probability_val)
        for i in range(num_simulations)
    )

    successful_results = [r for r in all_results if r is not None]

    if not successful_results:
        raise optuna.TrialPruned("All simulations failed for this parameter set.")

    results_df = pd.DataFrame(successful_results)
    output_cols = [col for col in results_df.columns if col not in input_df.index.to_list() and col != 'simulation_number']
    output_results_df = results_df[output_cols]

    if output_results_df.empty or len(output_results_df) < 2:
        return 0.0

    stats_df = output_results_df.describe().T
    
    if 'std' not in stats_df.columns or stats_df['std'].isnull().all():
        return 0.0

    avg_std = stats_df['std'].mean()
    
    return avg_std

def analyze_and_save_plots(study):
    """Generates and saves standard Optuna visualization plots."""

    if not study.trials:
        print("No trials to analyze.")
        return

    plots_dir = os.path.join('data', 'plots')
    os.makedirs(plots_dir, exist_ok=True)

    print("\nGenerating optimization analysis plots...")

    # Plot 1: Optimization History
    fig = optuna.visualization.plot_optimization_history(study)
    fig.write_html(os.path.join(plots_dir, "optimization_history.html"))

    # Plot 2: Parameter Importances
    completed_trials = [t for t in study.trials if t.state == optuna.trial.TrialState.COMPLETE]
    if len(completed_trials) > 1:
        try:
            fig = optuna.visualization.plot_param_importances(study)
            fig.write_html(os.path.join(plots_dir, "param_importances.html"))
        except Exception as e:
            # This can fail for various reasons, e.g., if all trials have the same value.
            print(f"Could not generate parameter importance plot: {e}")
    else:
        print("Skipping parameter importance plot (requires more than one completed trial).")
    
    print(f"Analysis plots saved as HTML files in the '{plots_dir}' directory.")

def run_bayesian_optimization(input_df, num_simulations_per_eval, max_calls, patience, probability_val):
    """
    Sets up and runs the Bayesian Optimization with Optuna.
    """
    optim_vars = input_df[input_df['randomizable'] == 1].index.tolist()
    
    objective = partial(objective_function,
                        base_input_df=input_df,
                        optim_vars=optim_vars,
                        num_simulations=num_simulations_per_eval,
                        probability_val=probability_val)

    print(f"\nStarting Bayesian Optimization for a maximum of {max_calls} trials...")
    print(f"Stopping if no improvement is seen for {patience} consecutive trials.")
    print(f"Optimizing default values for: {optim_vars}")
    print(f"Each trial will run {num_simulations_per_eval} stochastic simulations.")
    
    # Use TPE sampler, a form of Bayesian Optimization
    sampler = optuna.samplers.TPESampler(seed=42)
    study = optuna.create_study(direction='maximize', sampler=sampler)
    
    progress_callback = OptunaProgressCallback(patience=patience, max_calls=max_calls)
    
    interrupted = False
    initial_panel = Panel("Starting...", title="[bold green]Bayesian Optimization Status[/bold green]", border_style="green")
    with Live(initial_panel, screen=True, redirect_stderr=False, vertical_overflow="visible") as live:
        progress_callback.live = live
        try:
            # Optuna's warnings for repeated parameters can be verbose, so we hide them.
            optuna.logging.set_verbosity(optuna.logging.WARNING)
            study.optimize(
                objective,
                n_trials=max_calls,
                callbacks=[progress_callback],
                n_jobs=1  # Run trials sequentially; parallelization is inside the objective
            )
        except KeyboardInterrupt:
            interrupted = True
            print("\nOptimization stopped by user.")
        finally:
            optuna.logging.set_verbosity(optuna.logging.INFO) # Restore logging verbosity

    completed_trials = len(study.trials)
    if not interrupted and completed_trials < max_calls and study.best_trial:
         print("\nOptimization stopped early due to meeting the convergence criteria.")

    analyze_and_save_plots(study)

    try:
        best_params = study.best_params
        best_objective_value = study.best_value
    except ValueError: # This happens if no trials were completed successfully
        best_params = {}
        best_objective_value = None


    print("\nOptimization finished.")
    print(f"Total trials performed: {completed_trials}")
    if best_objective_value is not None:
        print(f"Best objective value (Maximized Avg. Std. Dev.): {best_objective_value:.4f}")
        print("Best default values found:")
        for var, val in best_params.items():
            print(f"  {var}: {val:.4f}")
    else:
        print("No successful trials completed.")
        return {}
        
    return best_params

# Main simulation function
def run_simulation():
    """
    Main function to drive the simulation based on user inputs and Excel data.
    """
    data_filepath = os.path.join('data', 'data.xlsx')

    print(f"Loading input variables from {data_filepath}...")
    try:
        input_df = pd.read_excel(data_filepath, sheet_name='input_config', index_col=0)
    except FileNotFoundError:
        print(f"Error: '{data_filepath}' not found. Please create it or run the script again to generate a template.")
        return
    except ValueError:
        print(f"Error: 'input_config' sheet not found in '{data_filepath}'.")
        return

    while True:
        choice = input("\nRun Bayesian Optimization to find best default values? (y/n): ").lower()
        if choice in ('y', 'n'):
            run_optimization = (choice == 'y')
            break
        print("Invalid choice. Please enter 'y' or 'n'.")

    num_simulations = 1
    vary_inputs = False
    probability_val = None

    if run_optimization:
        while True:
            try:
                max_calls = int(input("Enter the maximum number of optimization iterations: "))
                if max_calls > 0:
                    break
                print("Please enter a positive number.")
            except ValueError:
                print("Invalid input. Please enter an integer.")

        while True:
            try:
                patience = int(input("Enter convergence patience (stop after N iterations with no improvement): "))
                if patience > 0 and patience < max_calls:
                    break
                print(f"Patience must be a positive number less than the max iterations ({max_calls}).")
            except ValueError:
                print("Invalid input. Please enter an integer.")
        
        while True:
            try:
                n_sims_per_eval = int(input("Enter number of stochastic simulations per iteration: "))
                if n_sims_per_eval > 1:
                    break
                print("Please enter a number greater than 1 for meaningful statistics.")
            except ValueError:
                print("Invalid input. Please enter an integer.")
        
        while True:
            try:
                probability_val = float(input("Enter the probability for the confidence interval (e.g., 0.95): "))
                if 0 < probability_val < 1:
                    break
                print("Please enter a probability value between 0 and 1 (exclusive).")
            except ValueError:
                print("Invalid input. Please enter a floating-point number.")

        best_defaults = run_bayesian_optimization(input_df, n_sims_per_eval, max_calls, patience, probability_val)
        
        if best_defaults:
            print("\nUpdating default values in memory with optimized results...")
            for var, val in best_defaults.items():
                input_df.loc[var, 'default'] = val

            print(f"Saving optimal default values to '{data_filepath}'...")
            try:
                with pd.ExcelWriter(data_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    input_df.to_excel(writer, sheet_name='input_config')
                print("Optimal defaults saved successfully.")
            except Exception as e:
                print(f"An error occurred while saving the optimal defaults: {e}")
        else:
            print("Optimization did not yield a result. Using original default values for the final run.")

        print("\nNow, configure the final simulation run with the optimized default values.")
        while True:
            try:
                num_simulations = int(input("Enter the number of final simulations to run: "))
                if num_simulations > 0:
                    break
                print("Please enter a positive number.")
            except ValueError:
                print("Invalid input. Please enter an integer.")
        vary_inputs = True

    else:
        while True:
            try:
                num_simulations = int(input("Enter the number of simulations to run: "))
                if num_simulations > 0:
                    break
                print("Please enter a positive number.")
            except ValueError:
                print("Invalid input. Please enter an integer.")

        if num_simulations > 1:
            while True:
                choice = input("Vary inputs randomly? (y/n): ").lower()
                if choice in ('y', 'n'):
                    vary_inputs = (choice == 'y')
                    break
                print("Invalid choice. Please enter 'y' or 'n'.")
            
            if vary_inputs:
                while True:
                    try:
                        probability_val = float(input("Enter the probability for the confidence interval (e.g., 0.95): "))
                        if 0 < probability_val < 1:
                            break
                        print("Please enter a probability value between 0 and 1 (exclusive).")
                    except ValueError:
                        print("Invalid input. Please enter a floating-point number.")

    batch_size = 100
    total_successful = 0
    
    try:
        book = openpyxl.load_workbook(data_filepath)
        sheets_to_delete = ['results_statistics']
        for sheet_name in sheets_to_delete:
            if sheet_name in book.sheetnames:
                del book[sheet_name]
        book.save(data_filepath)
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"Warning: Could not clear previous result sheets in '{data_filepath}'. Error: {e}")

    print(f"\nStarting {num_simulations} simulations in batches of {batch_size}...")
    for i in tqdm(range(0, num_simulations, batch_size), desc="Processing Batches"):
        batch_end = min(i + batch_size, num_simulations)
        
        batch_results = Parallel(n_jobs=-1)(
            delayed(run_single_simulation)(sim_num, input_df, vary_inputs, probability_val)
            for sim_num in range(i, batch_end)
        )

        successful_batch_results = [r for r in batch_results if r is not None]
        total_successful += len(successful_batch_results)

        if not successful_batch_results:
            continue

        results_df = pd.DataFrame(successful_batch_results)
        
        input_cols = ['simulation_number'] + input_df.index.to_list()
        input_results_df = results_df.reindex(columns=input_cols)
        
        output_cols = ['simulation_number'] + [col for col in results_df.columns if col not in input_df.index.to_list() and col != 'simulation_number']
        output_results_df = results_df.reindex(columns=output_cols)

        try:
            with pd.ExcelWriter(data_filepath, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                is_new_input_sheet = 'results_input' not in writer.book.sheetnames
                input_start_row = 0 if is_new_input_sheet else writer.book['results_input'].max_row
                input_results_df.to_excel(writer, sheet_name='results_input', index=False, header=is_new_input_sheet, startrow=input_start_row)

                is_new_output_sheet = 'results_output' not in writer.book.sheetnames
                output_start_row = 0 if is_new_output_sheet else writer.book['results_output'].max_row
                output_results_df.to_excel(writer, sheet_name='results_output', index=False, header=is_new_output_sheet, startrow=output_start_row)
        except Exception as e:
            print(f"\nAn error occurred while saving batch results: {e}")
            break
        
        # Explicitly clear memory
        del batch_results
        del successful_batch_results
        del results_df
        del input_results_df
        del output_results_df
        gc.collect()

    print(f"\n{total_successful}/{num_simulations} simulations completed successfully.")

    if total_successful == 0:
        print("No results to save or analyze.")
        return

    print(f"All results have been incrementally saved to '{data_filepath}'.")
    
    if total_successful > 1:
        print("Calculating final statistics from the saved results...")
        try:
            final_output_df = pd.read_excel(data_filepath, sheet_name='results_output')
            stats_df = final_output_df.drop(columns=['simulation_number']).describe().T
            
            with pd.ExcelWriter(data_filepath, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                stats_df.to_excel(writer, sheet_name='results_statistics')
            
            print(f"Final statistics saved to sheet 'results_statistics' in '{data_filepath}'.")
        except Exception as e:
            print(f"An error occurred during final statistics calculation: {e}")

if __name__ == '__main__':
    run_simulation()